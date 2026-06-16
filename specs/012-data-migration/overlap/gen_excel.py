# -*- coding: utf-8 -*-
"""把"新旧系统重叠数据确认报告"导成多 sheet Excel（带颜色高亮 + 确认列）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "specs/012-data-migration/overlap/新旧系统重叠数据确认报告.xlsx"

# 样式
TITLE = Font(bold=True, size=14)
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="305496")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")      # 绿 一致
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")     # 红 撞车
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")    # 黄 差异
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD; cell.fill = HEAD_FILL; cell.alignment = CENTER; cell.border = BORDER

def apply_borders(ws, r1, r2, ncols):
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP

# ===== Sheet1 概览 =====
ws = wb.active; ws.title = "概览与结论"
ws["A1"] = "新旧系统重叠数据确认报告"; ws["A1"].font = TITLE
ws["A2"] = "数据来源：老库=yadapm_backup_20260611.dmp(Oracle实测)；新库=newpm生产K3s(备份newpm-20260615同步到ry_vue_prod)"
ws["A3"] = "生成日期：2026-06-15"
rows = [
    ["", ""],
    ["本次迁移涉及的 5 个功能", ""],
    ["① 批次版本管理", "出入库版本管理-批次版本"],
    ["② 非批次版本管理", "出入库版本管理-非批次版本"],
    ["③ 旧数据查询", "出入库版本管理-旧数据查询(历史归档)"],
    ["④ 批次任务问题单及缺陷", "项目质量管理-批次任务问题单及缺陷"],
    ["⑤ 非批次任务问题单及缺陷", "项目质量管理-非批次任务问题单及缺陷"],
    ["", ""],
    ["核心结论", ""],
    ["newpm 生产现状", "52 个任务(33个有任务号, 2025-2026) / 19 个批次(2026)；上述5功能业务表基本为空"],
    ["老 yadapm 历史数据", "跨 2021-2026，约 7200 条业务记录，引用约 2900 任务 / 218 批次"],
    ["任务重叠", "老『批次任务问题单及缺陷』引用202个去重任务，与生产仅重叠 7 个(其余195个生产没有)；其中2个为编号撞车"],
    ["批次重叠", "老库218个批次，与生产仅重叠 5 个(其余213个生产没有)"],
    ["一句话", "历史数据引用的任务/批次绝大多数在 newpm 不存在，迁移需业务方确认呈现策略"],
]
r = 4
for a, b in rows:
    ca = ws.cell(row=r, column=1, value=a); ca.font = Font(bold=True); ca.alignment = WRAP
    ws.cell(row=r, column=2, value=b).alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 90

# ===== Sheet2 任务号重叠 =====
ws = wb.create_sheet("任务号重叠(7)")
ws["A1"] = "任务号重叠（7个，需确认是否同一任务；同编号≠同任务）"; ws["A1"].font = TITLE
hdr = ["任务编号", "老系统任务名称", "newpm任务名称", "系统判断", "业务确认(请填)"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
task_rows = [
    ["M-202501-00030", "商户管理整合项目[第3期]", "商户管理整合项目[第3期]", "✅名称一致 疑似同一", ""],
    ["M-202510-00065", "网络支付整合与能力提升项目[第55期]", "TMS-【能力提升】-线下收单CRR086-273-201-中银智慧付静态二维码限额管理规则优化需求", "❌名称不同 疑似编号撞车", ""],
    ["M-202510-00632", "24年专项分期监管及审计整改项目[第10期]", "24年专项分期监管及审计整改项目[第10期]", "✅名称一致 疑似同一", ""],
    ["M-202511-00610", "网络支付整合与能力提升项目[第63期]", "TMS【能力提升】-线下收单CRR086-211-提升系统处理银联二维码反扫交易的成功率", "❌名称不同 疑似编号撞车", ""],
    ["M-202512-00694", "总对总汽车品牌购车线索管理及流程优化需求[第1期]", "总对总汽车品牌购车线索管理及流程优化需求[第1期]", "✅名称一致 疑似同一", ""],
    ["M-202601-00522", "商户管理整合项目[第19期]", "商户管理整合项目[第19期]", "✅名称一致 疑似同一", ""],
    ["M-202601-00645", "分行产品问题快速解决机制业务需求（中银智慧商家需求）[第4期]", "分行产品问题快速解决机制业务需求（中银智慧商家需求）[第4期]", "✅名称一致 疑似同一", ""],
]
for row in task_rows:
    ws.append(row)
apply_borders(ws, 4, 3 + len(task_rows), len(hdr))
for i, row in enumerate(task_rows):
    fill = BAD_FILL if "撞车" in row[3] else OK_FILL
    ws.cell(row=4 + i, column=4).fill = fill
widths = [18, 42, 50, 22, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# ===== Sheet3 批次号重叠 =====
ws = wb.create_sheet("批次号重叠(5)")
ws["A1"] = "批次号重叠（5个，需确认是否同一批次）"; ws["A1"].font = TITLE
hdr = ["批次号", "老系统计划投产日期", "newpm计划投产日期", "系统判断", "业务确认(请填)"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
batch_rows = [
    ["26年1月独立", "2026-01-23", "2026-01-23", "✅一致", ""],
    ["26年2月独立", "2026-02-12", "2026-02-12", "✅一致", ""],
    ["26年3月独立", "2026-03-27", "2026-03-27", "✅一致", ""],
    ["26年4月独立", "2026-04-17", "2026-04-17", "✅一致", ""],
    ["26年6月独立", "2026-06-25", "2026-06-26", "⚠️差1天 请确认以哪个为准", ""],
]
for row in batch_rows:
    ws.append(row)
apply_borders(ws, 4, 3 + len(batch_rows), len(hdr))
for i, row in enumerate(batch_rows):
    ws.cell(row=4 + i, column=4).fill = WARN_FILL if "差" in row[3] else OK_FILL
for i, w in enumerate([16, 20, 20, 26, 20], 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.append([]); ws.append(["注：老库另有 213 个历史批次(企架1.1批次/X62（C1601）等 2016-2025)，newpm 均无。"])

# ===== Sheet4 不重叠统计 =====
ws = wb.create_sheet("不重叠统计")
ws["A1"] = "不重叠部分（占绝大多数，迁移决策大头）"; ws["A1"].font = TITLE
hdr = ["维度", "老库总量", "与newpm重叠", "newpm没有", "占比"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
stat_rows = [
    ["批次任务问题单及缺陷-引用的任务", 202, 7, 195, "96.5%"],
    ["批次", 218, 5, 213, "97.7%"],
]
for row in stat_rows:
    ws.append(row)
apply_borders(ws, 4, 3 + len(stat_rows), len(hdr))
for i, w in enumerate([24, 12, 14, 12, 10], 1):
    ws.column_dimensions[chr(64 + i)].width = w

# ===== Sheet5 待确认问题 =====
ws = wb.create_sheet("待确认问题")
ws["A1"] = "需要相关人员确认/拍板的问题"; ws["A1"].font = TITLE
hdr = ["#", "问题", "业务答复(请填)"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
q_rows = [
    [1, "重叠的5个任务、5个批次：确认是否真为同一实体？(决定历史数据能否挂到现有任务/批次)", ""],
    [2, "2个编号撞车任务(M-202510-00065 / M-202511-00610)：怎么处理？", ""],
    [3, "6月批次差1天(老06-25 / 新06-26)：以哪边为准？", ""],
    [4, "占97%的『newpm没有的历史任务/批次』怎么呈现？ A反范式化(历史记录自带任务名/批次号文本快照,推荐) / B注入老任务批次(把老任务批次也导进系统,会污染日常任务管理) / C缩范围(先只迁自包含的功能:旧数据查询、非批次版本管理、非批次任务问题单及缺陷)", ""],
    [5, "附件：老『批次任务问题单及缺陷』仅47个附件、『非批次任务问题单及缺陷』0个，且物理文件不在备份(只有DB记录指向老服务器F:\\路径)。是否需单独找老服务器要附件文件？", ""],
]
for row in q_rows:
    ws.append(row)
apply_borders(ws, 4, 3 + len(q_rows), len(hdr))
for i, w in enumerate([5, 80, 30], 1):
    ws.column_dimensions[chr(64 + i)].width = w

wb.save(OUT)
print("已生成:", OUT)
